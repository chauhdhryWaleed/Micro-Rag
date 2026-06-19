"""
Ingest: XLSX → OpenAI embeddings → Qdrant

Design decisions worth noting:

1. We embed a curated text blob per record, not the whole row.
   Embedding all 43 fields would dilute the signal — an email address
   or pipeline_run_date has no semantic value for retrieval. We embed
   only the fields that describe what the office *does*, then store
   everything else as payload for metadata filtering and display.

2. We batch all embeddings in a single OpenAI call. At 50 records,
   this is ~5K tokens — well under the 300K token batch limit and
   costs roughly $0.0001. Not worth optimising further.

3. Type coercion before upsert: Qdrant range filters only work on
   numeric payload values. Fields like confidence_score arrive as
   floats from Claude, but Excel may serialise them as strings.
   We coerce to the right type or drop the value so filters don't
   silently break. [verified: Qdrant rejects string values in Range conditions]

4. We add a synthetic `_embed_text` field to the payload so the
   retrieval layer can show users what text drove the similarity score.
   Useful for debugging and for building user trust — "here is exactly
   what we matched against" is more honest than a black-box score.
"""

import logging
import sys
from pathlib import Path
from openai import OpenAI
from qdrant_client import QdrantClient

logger = logging.getLogger(__name__)
from qdrant_client.models import (
    Distance, VectorParams, PointStruct, PayloadSchemaType
)
from openpyxl import load_workbook

# Allow running as `python -m rag.ingest` from project root
sys.path.insert(0, str(Path(__file__).parent.parent))

from pipeline.config import COLUMN_HEADERS, SCHEMA_FIELDS
from rag.config import (
    OPENAI_API_KEY, QDRANT_URL, QDRANT_API_KEY,
    COLLECTION_NAME, EMBEDDING_MODEL, EMBEDDING_DIM,
    SEMANTIC_FIELDS, XLSX_PATH,
)


def get_qdrant_client() -> QdrantClient:
    if QDRANT_URL:
        return QdrantClient(
            url=QDRANT_URL,
            api_key=QDRANT_API_KEY or None,
            check_compatibility=False,  # suppresses version-check warning on Qdrant Cloud
        )
    return QdrantClient(":memory:")


def build_embed_text(record: dict) -> str:
    """Concatenate semantic fields into one string for embedding."""
    parts = []
    for field in SEMANTIC_FIELDS:
        val = record.get(field)
        if val and str(val).strip().lower() not in ("", "null", "none", "n/a"):
            parts.append(str(val).strip())
    return " | ".join(parts)


def coerce_types(record: dict) -> dict:
    """
    Cast numeric fields to their correct Python types before storing as Qdrant payload.
    If coercion fails (e.g., Excel stored "N/A" in confidence_score), we set None
    so the field is absent from the payload rather than breaking range filter queries.
    """
    int_fields   = ["founding_year", "aum_year", "data_completion_score"]
    float_fields = ["confidence_score"]

    for f in int_fields:
        try:
            record[f] = int(record[f]) if record.get(f) not in (None, "") else None
        except (ValueError, TypeError):
            record[f] = None

    for f in float_fields:
        try:
            record[f] = float(record[f]) if record.get(f) not in (None, "") else None
        except (ValueError, TypeError):
            record[f] = None

    return record


def load_records_from_xlsx(path: str) -> list[dict]:
    """
    Read the pipeline-generated XLSX.
    Row 1 = section header bands (skip)
    Row 2 = column headers → map to field names via reverse COLUMN_HEADERS lookup
    Row 3+ = data

    We reverse-lookup field names from display headers rather than assuming column order,
    so the ingest survives if someone manually reorders columns in the spreadsheet.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_to_field = {v: k for k, v in COLUMN_HEADERS.items()}

    # Row 2 contains column headers
    col_headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    field_names = [header_to_field.get(h) for h in col_headers]

    records = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if all(v is None for v in row):
            continue
        record = {}
        for field, val in zip(field_names, row):
            if field:
                record[field] = val if val is not None else None
        records.append(record)

    logger.info("Loaded %d records from %s", len(records), path)
    return records


def ingest(xlsx_path: str = XLSX_PATH, client: QdrantClient = None) -> int:
    """
    Full ingest pipeline: XLSX → embed → upsert to Qdrant.
    Returns number of records indexed.

    Validation check at the end confirms Qdrant's count matches what we loaded.
    A mismatch here would indicate a silent partial failure — we surface it rather
    than letting the API serve incomplete data with false confidence.
    """
    if client is None:
        client = get_qdrant_client()

    oai = OpenAI(api_key=OPENAI_API_KEY)

    # --- Load ---
    records = load_records_from_xlsx(xlsx_path)
    if not records:
        logger.warning("No records found at %s — run the data collection pipeline first", xlsx_path)
        return 0

    # --- Coerce types ---
    records = [coerce_types(r) for r in records]

    # --- Build embedding texts ---
    embed_texts = [build_embed_text(r) for r in records]
    for r, t in zip(records, embed_texts):
        r["_embed_text"] = t  # store for transparency/debugging

    # --- Skip re-ingest if collection already has the right record count ---
    # In-memory Qdrant: collection is always empty on startup, always ingest.
    # Qdrant Cloud: collection persists across restarts — skip embedding if count matches.
    existing = [c.name for c in client.get_collections().collections]
    if COLLECTION_NAME in existing:
        current_count = client.count(collection_name=COLLECTION_NAME).count
        if current_count == len(records):
            logger.info(
                "Collection '%s' already contains %d records — skipping re-ingest. [verified]",
                COLLECTION_NAME, current_count,
            )
            return current_count
        logger.info(
            "Collection count mismatch (%d in Qdrant vs %d in XLSX) — rebuilding.",
            current_count, len(records),
        )
        client.delete_collection(COLLECTION_NAME)

    logger.info("Embedding %d records with %s...", len(records), EMBEDDING_MODEL)
    response = oai.embeddings.create(model=EMBEDDING_MODEL, input=embed_texts)
    vectors = [item.embedding for item in response.data]

    client.create_collection(
        collection_name=COLLECTION_NAME,
        vectors_config=VectorParams(size=EMBEDDING_DIM, distance=Distance.COSINE),
    )

    # --- Upsert ---
    points = [
        PointStruct(id=i, vector=vec, payload=rec)
        for i, (vec, rec) in enumerate(zip(vectors, records))
    ]
    client.upsert(collection_name=COLLECTION_NAME, points=points)

    # --- Validate: count in Qdrant must match what we loaded ---
    indexed = client.count(collection_name=COLLECTION_NAME).count
    if indexed != len(records):
        logger.warning(
            "Count mismatch: loaded %d records but Qdrant reports %d. "
            "Partial ingest — do not trust retrieval results. [inferred]",
            len(records), indexed,
        )
    else:
        logger.info("%d records indexed and count verified. [verified]", indexed)

    return indexed


if __name__ == "__main__":
    from logging_config import setup_logging
    setup_logging()
    client = get_qdrant_client()
    n = ingest(client=client)
    logger.info("Done. %d records ready for search.", n)
