"""
Phase 4: Export

Writes the validated dataset to XLSX and a methodology JSON log.
XLSX formatting mirrors the sample FO-MAX-data-sample-2.0.xlsx structure.
"""

import json
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pipeline.config import SCHEMA_FIELDS, COLUMN_HEADERS, XLSX_PATH, METHODOLOGY_PATH


# Color scheme matching the sample file's professional look
HEADER_BG = "1F3864"   # dark navy
HEADER_FG = "FFFFFF"
SECTION_COLORS = {
    "entity":     "2E75B6",  # blue
    "principal":  "375623",  # dark green
    "signals":    "7030A0",  # purple
    "meta":       "404040",  # dark grey
}

SECTION_RANGES = {
    "entity":    slice(0, 17),
    "principal": slice(17, 26),
    "signals":   slice(26, 36),
    "meta":      slice(36, None),
}


def _section_for_col(col_idx: int) -> str:
    for name, slc in SECTION_RANGES.items():
        start = slc.start or 0
        stop = slc.stop or len(SCHEMA_FIELDS)
        if start <= col_idx < stop:
            return name
    return "meta"


def _thin_border():
    thin = Side(border_style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def export_xlsx(records: list[dict]) -> str:
    """Write records to XLSX. Returns path to the output file."""
    os.makedirs(os.path.dirname(XLSX_PATH), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Family Offices"

    # ---- Section header row (row 1) ----
    section_labels = {
        "entity": "ENTITY ATTRIBUTES",
        "principal": "PRINCIPAL INTELLIGENCE",
        "signals": "ENTITY SIGNALS / RECENT ACTIVITIES",
        "meta": "VALIDATION & META",
    }
    section_start_cols = {}
    for name, slc in SECTION_RANGES.items():
        start = (slc.start or 0) + 1   # 1-indexed, +1 for Excel
        section_start_cols[name] = start

    for section_name, start_col_0idx in [
        ("entity", 0), ("principal", 17), ("signals", 26), ("meta", 36)
    ]:
        slc = SECTION_RANGES[section_name]
        s = slc.start or 0
        e = slc.stop or len(SCHEMA_FIELDS)
        start_col = s + 1
        end_col = e

        cell = ws.cell(row=1, column=start_col, value=section_labels[section_name])
        color = SECTION_COLORS[section_name]
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if end_col > start_col:
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col
            )

    # ---- Column header row (row 2) ----
    for col_idx, field in enumerate(SCHEMA_FIELDS):
        col = col_idx + 1
        header = COLUMN_HEADERS.get(field, field)
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()

    # ---- Data rows (starting row 3) ----
    for row_idx, record in enumerate(records):
        row = row_idx + 3
        for col_idx, field in enumerate(SCHEMA_FIELDS):
            col = col_idx + 1
            val = record.get(field)
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            if val is None or str(val).strip() in ("null", "None", ""):
                val = ""

            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _thin_border()
            cell.font = Font(size=9)

            # Zebra striping
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F2F2F2")

    # ---- Column widths ----
    col_widths = {
        "fo_name": 30, "description": 45, "investment_thesis": 45,
        "investment_mandate": 35, "aum_estimate": 18, "corporate_linkedin": 35,
        "website_url": 30, "asset_classes": 30, "geographic_focus": 25,
        "principal_full_name": 22, "principal_title": 25, "principal_linkedin": 35,
        "principal_email": 28, "recent_news_headline": 40, "enrichment_sources": 40,
    }
    default_width = 16
    for col_idx, field in enumerate(SCHEMA_FIELDS):
        col_letter = get_column_letter(col_idx + 1)
        ws.column_dimensions[col_letter].width = col_widths.get(field, default_width)

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"

    wb.save(XLSX_PATH)
    print(f"[Export] Saved {len(records)} records to {XLSX_PATH}")
    return XLSX_PATH


def export_methodology(
    records: list[dict],
    discovery_log: dict,
    validation_summary: dict,
) -> str:
    """Write a methodology JSON log for submission."""
    os.makedirs(os.path.dirname(METHODOLOGY_PATH), exist_ok=True)

    # Pick the 3 highest-confidence records for full validation chains
    top3 = sorted(records, key=lambda r: r.get("confidence_score", 0), reverse=True)[:3]

    methodology = {
        "run_date": str(date.today()),
        "pipeline_version": "1.0",
        "overview": {
            "discovery_sources": ["SEC EDGAR Form ADV", "Gemini 2.5 Flash-Lite web search"],
            "enrichment_method": "Claude claude-sonnet-4-6 tool-use agent (search_web + scrape_url)",
            "validation_method": "URL HTTP check + DNS MX record validation + field presence checks",
            "total_candidates_discovered": discovery_log.get("total_candidates", 0),
            "records_enriched": discovery_log.get("records_enriched", 0),
            "records_passing_validation": len(records),
            "target_records": 50,
        },
        "schema_decisions": {
            "entity_attributes": "17 fields covering entity identity, mandate, and online presence",
            "principal_intelligence": "9 fields; email inferred via pattern matching, validated via MX DNS",
            "entity_signals": "10 fields; sourced from web news search, last 12-24 months",
            "confidence_scoring": "1-10 scale: starts from Claude self-report, adjusted by validation passes/failures",
        },
        "known_limitations": [
            "AUM figures are often self-reported or estimated from news; not independently audited",
            "Principal emails are mostly pattern-inferred (first@domain), not confirmed via SMTP",
            "Some family offices deliberately maintain low digital footprints — discovery is biased toward visible FOs",
            "SEC EDGAR covers only registered advisers (>$100M AUM threshold); smaller SFOs may be underrepresented",
            "LinkedIn URLs require manual verification — the pipeline validates URL format only, not actual page existence",
        ],
        "validation_summary": validation_summary,
        "full_validation_chains": [
            _build_validation_chain(r) for r in top3
        ],
    }

    with open(METHODOLOGY_PATH, "w") as f:
        json.dump(methodology, f, indent=2, ensure_ascii=False)

    print(f"[Export] Methodology log saved to {METHODOLOGY_PATH}")
    return METHODOLOGY_PATH


def _build_validation_chain(record: dict) -> dict:
    """Build a full validation chain for a single record (for methodology doc)."""
    return {
        "fo_name": record.get("fo_name", ""),
        "discovery": {
            "source": record.get("discovery_source", ""),
            "url": record.get("discovery_url", ""),
        },
        "extraction_method": "Claude claude-sonnet-4-6 tool-use agent: web search + URL scraping",
        "enrichment_steps": [
            "Broad web search for entity + 'family office'",
            "Targeted search for AUM and investment thesis",
            "Principal search: '[Name] CEO/CIO/Managing Director'",
            "News search: '[Name] 2025 2026 investment hire'",
            "Website scrape for primary source data",
        ],
        "validation_logic": {
            "website_check": "HTTP HEAD request, verified reachable",
            "email_check": "DNS MX record lookup on email domain",
            "linkedin_check": "URL format pattern match only",
            "field_presence": "Minimum required fields checked",
        },
        "confidence_assessment": {
            "score": record.get("confidence_score", 0),
            "validation_status": record.get("validation_status", ""),
            "data_completion_pct": record.get("data_completion_score", 0),
            "validations_passed": record.get("_validations_passed", []),
            "issues_found": record.get("_validation_issues", []),
        },
        "enrichment_sources": record.get("enrichment_sources", ""),
    }
